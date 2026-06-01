"""
exportacao_fiscal.py — Exportação Excel para Fiscalização UFAC
Formato: aba única "Relatorio Fiscal", Etapa como coluna (não como aba).
"""
import io
from datetime import date
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Constantes de estilo ───────────────────────────────────
AZUL_ESCURO = "1F4E78"
AZUL_CLARO = "D6E4F0"
CINZA_ZEBRA = "F2F2F2"
BRANCO = "FFFFFF"
BORDA_COR = "BFBFBF"

# ── Mapeamento de colunas (origem → destino) ───────────────
MAPA_COLUNAS = {
    "ID": "O.S.",
    "Titulo": "Título",
    "Contrato": "Contrato",
    "Etapa": "Etapa",
    "Fiscal responsavel": "Fiscal Responsável",
    "Valor total": "Valor Total (R$)",
    "Valor M.O": "Mão de Obra (R$)",
    "Valor M.A": "Material (R$)",
    "Local do servico": "Local",
    "Descricao previa": "Descrição Prévia",
}

# Coluna opcional (só aparece se existir no df)
COLUNA_DESCRICAO_DETALHADA = "Descricao detalhada da solicitacao"

# Larguras fixas por coluna de destino
LARGURAS = {
    "O.S.": 10,
    "Título": 60,
    "Contrato": 12,
    "Etapa": 38,
    "Fiscal Responsável": 28,
    "Valor Total (R$)": 20,
    "Mão de Obra (R$)": 20,
    "Material (R$)": 18,
    "Local": 40,
    "Descrição Prévia": 35,
    "Descrição Detalhada": 50,
}

COLUNAS_VALOR = {"Valor Total (R$)", "Mão de Obra (R$)", "Material (R$)"}


# ── Helpers ─────────────────────────────────────────────────

def _to_float(v) -> float:
    """Converte valor brasileiro '1.234,56' para float."""
    if v is None:
        return 0.0
    s = str(v).strip()
    if s in ("", "-", "#N/A", "0", "None"):
        return 0.0
    try:
        return float(s.replace(".", "").replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def _borda() -> Border:
    s = Side(style="thin", color=BORDA_COR)
    return Border(left=s, right=s, top=s, bottom=s)


def _cabecalho(ws, row: int, nc: int):
    """Aplica estilo de cabeçalho na linha: fundo azul escuro, fonte branca."""
    for c in range(1, nc + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color=BRANCO, name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = _borda()


# ── Preparação do DataFrame ─────────────────────────────────

def _preparar_df(df_filtrado: pd.DataFrame) -> pd.DataFrame:
    """
    Seleciona e renomeia colunas conforme MAPA_COLUNAS.
    Inclui 'Descrição Detalhada' se a coluna existir no df original.
    Trabalha em cópia — não modifica o original.
    """
    # Constrói mapa dinâmico (inclui descrição detalhada se existir)
    mapa = dict(MAPA_COLUNAS)
    if COLUNA_DESCRICAO_DETALHADA in df_filtrado.columns:
        mapa[COLUNA_DESCRICAO_DETALHADA] = "Descrição Detalhada"

    # Só seleciona colunas que existem
    colunas_existentes = {k: v for k, v in mapa.items() if k in df_filtrado.columns}
    out = df_filtrado[list(colunas_existentes.keys())].copy()
    out.rename(columns=colunas_existentes, inplace=True)

    # Converte valores monetários
    for c in COLUNAS_VALOR:
        if c in out.columns:
            out[c] = out[c].apply(_to_float)

    return out


# ── Geração do Excel ────────────────────────────────────────

def _gerar_excel(df_prep: pd.DataFrame, contratos: list, etapas: list) -> bytes:
    """
    Gera .xlsx em memória com 1 aba 'Relatorio Fiscal'.
    Retorna bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio Fiscal"
    ws.sheet_view.showGridLines = False

    nc = len(df_prep.columns)
    ultima_col = get_column_letter(nc)

    # ── Linha 1: Título ──
    ws.merge_cells(f"A1:{ultima_col}1")
    contratos_str = ", ".join(contratos)
    data_str = date.today().strftime("%d/%m/%Y")
    ws["A1"] = f"RELATÓRIO DE FISCALIZAÇÃO — UFAC | Contratos: {contratos_str} | {data_str}"
    ws["A1"].font = Font(bold=True, size=13, color=BRANCO, name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Linha 2: Sub-header com totais ──
    ws.merge_cells(f"A2:{ultima_col}2")
    n_os = len(df_prep)
    vt = df_prep["Valor Total (R$)"].sum() if "Valor Total (R$)" in df_prep.columns else 0
    mo = df_prep["Mão de Obra (R$)"].sum() if "Mão de Obra (R$)" in df_prep.columns else 0
    ma = df_prep["Material (R$)"].sum() if "Material (R$)" in df_prep.columns else 0
    ws["A2"] = (
        f"Total: {n_os} O.S.  |  Valor Total: R$ {vt:,.2f}  |  "
        f"Mão de Obra: R$ {mo:,.2f}  |  Material: R$ {ma:,.2f}"
    )
    ws["A2"].font = Font(italic=True, size=10, color="404040", name="Calibri")
    ws["A2"].fill = PatternFill("solid", fgColor=AZUL_CLARO)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Linha 3: Cabeçalho das colunas ──
    HDR_ROW = 3
    for ci, col_name in enumerate(df_prep.columns, 1):
        ws.cell(HDR_ROW, ci, col_name)
    _cabecalho(ws, HDR_ROW, nc)
    ws.row_dimensions[HDR_ROW].height = 22

    # ── Linhas 4+: Dados ──
    for ri, row in enumerate(df_prep.itertuples(index=False), HDR_ROW + 1):
        for ci, val in enumerate(row, 1):
            col_name = df_prep.columns[ci - 1]
            safe = "" if (not isinstance(val, str) and pd.isna(val)) else val
            cell = ws.cell(ri, ci, safe)

            # Zebra striping
            cor = CINZA_ZEBRA if ri % 2 == 0 else BRANCO
            cell.fill = PatternFill("solid", fgColor=cor)
            cell.border = _borda()
            cell.font = Font(name="Calibri", size=10)

            # Formatação por tipo
            if col_name in COLUNAS_VALOR:
                cell.number_format = "R$ #,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name == "O.S.":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in ("Contrato",):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "Fiscal Responsável":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        ws.row_dimensions[ri].height = 18

    # ── Larguras fixas ──
    for ci, col_name in enumerate(df_prep.columns, 1):
        largura = LARGURAS.get(col_name, 20)
        ws.column_dimensions[get_column_letter(ci)].width = largura

    # ── Freeze + AutoFilter ──
    ws.freeze_panes = f"A{HDR_ROW + 1}"
    ws.auto_filter.ref = f"A{HDR_ROW}:{ultima_col}{HDR_ROW + len(df_prep)}"

    # ── Salva em memória ──
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Componente Streamlit ────────────────────────────────────

def render_exportacao_fiscal(df_ufac: pd.DataFrame) -> None:
    """
    Renderiza painel de exportação no Streamlit.
    Plug-in: from exportacao_fiscal import render_exportacao_fiscal
    Uso: render_exportacao_fiscal(df_ufac)
    """
    with st.expander("📥 Exportar Relatório para Fiscalização", expanded=False):
        st.markdown("##### Configurar exportação")

        c1, c2 = st.columns(2)

        with c1:
            contratos_disponiveis = (
                sorted(df_ufac["Contrato"].dropna().unique().tolist())
                if "Contrato" in df_ufac.columns
                else []
            )
            contratos_sel = st.multiselect(
                "Contrato(s)",
                options=contratos_disponiveis,
                default=contratos_disponiveis,
            )

        with c2:
            etapas_disponiveis = (
                sorted(df_ufac["Etapa"].dropna().unique().tolist())
                if "Etapa" in df_ufac.columns
                else []
            )
            etapas_sel = st.multiselect(
                "Etapa(s)",
                options=etapas_disponiveis,
                default=etapas_disponiveis,
            )

        if not contratos_sel:
            st.warning("Selecione ao menos um contrato.")
            return
        if not etapas_sel:
            st.warning("Selecione ao menos uma etapa.")
            return

        # Filtra
        df2 = df_ufac[
            df_ufac["Contrato"].isin(contratos_sel) & df_ufac["Etapa"].isin(etapas_sel)
        ]

        total = len(df2)
        st.info(
            f"ℹ️ **{total} O.S. encontradas** | "
            f"Contratos: {', '.join(contratos_sel)} | "
            f"{len(etapas_sel)} etapa{'s' if len(etapas_sel) != 1 else ''} selecionada{'s' if len(etapas_sel) != 1 else ''}"
        )

        if total == 0:
            st.error("Nenhuma O.S. encontrada com os filtros atuais.")
            return

        # Prepara e gera Excel
        df_prep = _preparar_df(df2)
        xlsx_bytes = _gerar_excel(df_prep, contratos_sel, etapas_sel)

        # Nome do arquivo
        contratos_clean = "_".join(contratos_sel).replace("/", "")
        nome_arquivo = f"Relatorio_Fiscal_UFAC_{contratos_clean}_{date.today().strftime('%Y%m%d')}.xlsx"

        st.download_button(
            label="⬇️  Baixar Relatório Excel",
            data=xlsx_bytes,
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
